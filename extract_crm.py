#!/usr/bin/env python3
import json
import openpyxl

wb = openpyxl.load_workbook("/Users/kd.macbot/.openclaw/workspace/consultplus-crm/CRM data.xlsx", data_only=True)

result = {}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        result[sheet_name] = {"headers": [], "data": []}
        continue

    # First row = headers
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    print(f"=== Sheet: {sheet_name} === ({len(rows)-1} data rows)")
    print(f"Headers: {headers}")
    print()

    data = []
    for row in rows[1:]:
        record = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            # Convert dates/times to string for JSON
            if hasattr(val, 'isoformat'):
                val = val.isoformat()
            record[key] = val
        # Skip completely empty rows
        if any(v is not None for v in record.values()):
            data.append(record)

    result[sheet_name] = {"headers": headers, "row_count": len(data), "data": data}

print("=" * 60)
print("FULL JSON OUTPUT:")
print("=" * 60)
print(json.dumps(result, indent=2, default=str, ensure_ascii=False))
